from openpyxl.styles import Font, Alignment, PatternFill

class ExcelStyle:
    def __init__(self, sheet):
        """
        Recebe uma aba (worksheet) do openpyxl para aplicar estilos.
        """
        self.sheet = sheet

    def format_cell(self, row, col, bold=False, align_center=False, bg_color=None):
        """Aplica formatação em uma célula."""
        cell = self.sheet.cell(row=row, column=col)
        if bold:
            cell.font = Font(bold=True)
        if align_center:
            cell.alignment = Alignment(horizontal="center")
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        return cell

    def format_row(self, row_num, bold=False, align_center=False, bg_color=None):
        """Formata uma linha inteira."""
        for col in range(1, self.sheet.max_column + 1):
            self.format_cell(row_num, col, bold=bold, align_center=align_center, bg_color=bg_color)

    def autosize_columns(self):
        """Ajusta automaticamente a largura das colunas da aba ativa."""
        for col in self.sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            self.sheet.column_dimensions[col_letter].width = max_length + 2
