from openpyxl import Workbook, load_workbook

class ExcelCore:
    def __init__(self, filename=None):
        """
        Se filename for fornecido, abre o arquivo.
        Caso contrário, cria um novo workbook.
        """
        if filename:
            self.workbook = load_workbook(filename)
            self.filename = filename
        else:
            self.workbook = Workbook()
            self.filename = None

        self.sheet = self.workbook.active

    def set_active_sheet(self, name):
        """Seleciona a aba pelo nome (se existir)."""
        if name in self.workbook.sheetnames:
            self.sheet = self.workbook[name]
        else:
            raise ValueError(f"Aba '{name}' não existe")

    def create_sheet(self, name):
        """Cria uma nova aba e a define como ativa."""
        self.sheet = self.workbook.create_sheet(title=name)
        return self.sheet

    def write_cell(self, row, col, value):
        """Escreve valor em uma célula (sem formatação)."""
        self.sheet.cell(row=row, column=col, value=value)

    def write_row(self, row_num, values):
        """Escreve uma lista de valores em uma linha inteira (sem formatação)."""
        for col, value in enumerate(values, start=1):
            self.write_cell(row_num, col, value)

    def read_row(self, row_num):
        """Lê os valores de uma linha."""
        return [cell.value for cell in self.sheet[row_num]]

    def save(self, filename=None):
        """Salva o arquivo Excel."""
        if filename:
            self.filename = filename
        if not self.filename:
            raise ValueError("Nenhum nome de arquivo definido para salvar")
        self.workbook.save(self.filename)
